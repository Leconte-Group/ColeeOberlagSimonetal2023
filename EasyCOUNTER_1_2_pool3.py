#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Wed Jul 26 14:47:12 2023

@author: aaronmleconte
"""

#In this program, I am trying to take EasyDIVER files, translate the 
#sequences, and then pull out the number of mutations in each.  


#Start by inputing a file name

filename="MS2_078_pool3.xlsx"
inputsheet="Sheet1"

#Open program that works with Excel and set the sheet1 as the active sheet
import openpyxl
wb=openpyxl.load_workbook(filename)
type(wb)
sheet1=wb[inputsheet]

#Input the barcodes that will be used.  These should be the reverse
#complement of the primer sequence.
bar_D_bri="CGTGATCTAT" #528
bar_B_bri="GATCTGCTAT" #529
bar_D_sel="ACATCGCTAT" #530
bar_B_sel="TCAAGTCTAT" #531
bar_active="GCCTAACTAT" #532

#this is the amplicon without the forward primer
amplicon="CCCATCTTCGGCAACCAGATCATCCCCGACACCGCTATCCTCAGCGTGGTGCCATTTCACCACGGCTTCGGCATGTTCACCACGCTGGGCTACTTGATCTGCGGCTTTCGGGTCGTGCTCATGTACCGCTTCGAGGAGGAGCTATTCTTGCGaAGCGTGATCTAT"

#if using a fourth misc barcode, then add it to this list below
barcode_list=[bar_D_bri, bar_B_bri, bar_D_sel, bar_B_sel, bar_active]
barcode_names=["bar_D_bri", "bar_B_bri", "bar_D_sel", "bar_B_sel", "bar_active"]

amino_acids_in_library = ["243","246","250"]

amino_acid_list = ['A','C','D','E','F','G','H','I','K','L','M','N','P','Q','R','S','T','V','W','X','Y','_']

#code to translate a sequence into a protein sequence
def translate(seq):
	
    table = {
		'ATA':'I', 'ATC':'I', 'ATT':'I', 'ATG':'M',
		'ACA':'T', 'ACC':'T', 'ACG':'T', 'ACT':'T',
		'AAC':'N', 'AAT':'N', 'AAA':'K', 'AAG':'K',
		'AGC':'S', 'AGT':'S', 'AGA':'R', 'AGG':'R',				
		'CTA':'L', 'CTC':'L', 'CTG':'L', 'CTT':'L',
		'CCA':'P', 'CCC':'P', 'CCG':'P', 'CCT':'P',
		'CAC':'H', 'CAT':'H', 'CAA':'Q', 'CAG':'Q',
		'CGA':'R', 'CGC':'R', 'CGG':'R', 'CGT':'R',
		'GTA':'V', 'GTC':'V', 'GTG':'V', 'GTT':'V',
		'GCA':'A', 'GCC':'A', 'GCG':'A', 'GCT':'A',
		'GAC':'D', 'GAT':'D', 'GAA':'E', 'GAG':'E',
		'GGA':'G', 'GGC':'G', 'GGG':'G', 'GGT':'G',
		'TCA':'S', 'TCC':'S', 'TCG':'S', 'TCT':'S',
		'TTC':'F', 'TTT':'F', 'TTA':'L', 'TTG':'L',
		'TAC':'Y', 'TAT':'Y', 'TAA':'_', 'TAG':'_',
		'TGC':'C', 'TGT':'C', 'TGA':'_', 'TGG':'W',
        }
    protein =""
    if len(seq)%3 == 0:
        for i in range(0, len(seq), 3):
            codon = seq[i:i+3]
            if "N" in codon:
                protein = protein + "X"
            else:
                protein = protein + table[codon]
    elif len(seq)%3 == 1:
        seq=seq[:-1]
        for i in range(0, len(seq), 3):
            codon = seq[i:i+3]
            if "N" in codon:
                protein = protein + "X"
            else:
                protein = protein + table[codon]
    elif len(seq)%3 == 2:
        seq=seq[:-2]
        for i in range(0, len(seq), 3):
            codon = seq[i:i+3]
            if "N" in codon:
                protein = protein + "X"
            else:
                protein = protein + table[codon]
    return protein


#start by creating a new list of sequences that are around the right length of 
#the amplicon (removes truncated PCR products) and then sorting

def sort_good_sequences_by_barcode(sheet):
#create the output workbook with a sheet for each barcode
    output=openpyxl.Workbook()
    for i in range(len(barcode_list)):
        output.create_sheet(barcode_names[i])
        for j in range(3,3+len(amino_acid_list)):
            output[barcode_names[i]].cell(j,1).value=amino_acid_list[j-3]
            for k in range(len(amino_acids_in_library)):
                output[barcode_names[i]].cell(1,k+2).value = amino_acids_in_library[k]
                output[barcode_names[i]].cell(j,k+2).value = 0
        
#remove short sequences
    output.create_sheet("Good Calls")
    for row in range(4,sheet1.max_row+1):
        sequence=str(sheet1.cell(row,1).value)
        if len(sequence)>(len(amplicon)-10):
            if len(sequence)<(len(amplicon)+10):
                for i in range(1,4):
                    output["Good Calls"].cell(row,i).value=sheet1.cell(row,i).value
 #identify the bar code sequence and write that into spreadsheet in column D
            for i in range(0,len(barcode_list)):
                if barcode_list[i] in sequence:
                    output["Good Calls"].cell(row,4).value=barcode_names[i]
#translate the sequence and print it in column E
    for row in range(4,sheet1.max_row+1):
        sequence=output["Good Calls"].cell(row,1).value
        print(sequence)
        if type(sequence)==str:
            protein = translate(sequence)
            print(protein)
            output["Good Calls"].cell(row,5).value=protein
            output["Good Calls"].cell(row,6).value=len(protein)
            output["Good Calls"].cell(row,6).value=protein[243-225]
            output["Good Calls"].cell(row,7).value=protein[246-225]
            output["Good Calls"].cell(row,8).value=protein[250-225]
#match mutations at position 1
    for row in range(4,output["Good Calls"].max_row+1):
        ind_seq_barcode=output["Good Calls"].cell(row,4).value
        if type(ind_seq_barcode) == str:
            for i in range(0,len(barcode_names)):
                if barcode_names[i]==ind_seq_barcode:
                    for row_bar in range(3,24):
                        if output["Good Calls"].cell(row,6).value==output[barcode_names[i]].cell(row_bar,1).value:
                            output[barcode_names[i]].cell(row_bar,2).value=output[barcode_names[i]].cell(row_bar,2).value+output["Good Calls"].cell(row,2).value
    for row in range(4,output["Good Calls"].max_row+1):
        ind_seq_barcode=output["Good Calls"].cell(row,4).value
        if type(ind_seq_barcode) == str:
            for i in range(0,len(barcode_names)):
               if barcode_names[i]==ind_seq_barcode:
                    for row_bar in range(3,24):
                        if output["Good Calls"].cell(row,7).value==output[barcode_names[i]].cell(row_bar,1).value:
                            output[barcode_names[i]].cell(row_bar,3).value=output[barcode_names[i]].cell(row_bar,3).value+output["Good Calls"].cell(row,2).value
    for row in range(4,output["Good Calls"].max_row+1):
        ind_seq_barcode=output["Good Calls"].cell(row,4).value
        if type(ind_seq_barcode) == str:
            for i in range(0,len(barcode_names)):
               if barcode_names[i]==ind_seq_barcode:
                    for row_bar in range(3,24):
                        if output["Good Calls"].cell(row,8).value==output[barcode_names[i]].cell(row_bar,1).value:
                            output[barcode_names[i]].cell(row_bar,4).value=output[barcode_names[i]].cell(row_bar,4).value+output["Good Calls"].cell(row,2).value
    print("All done!!")


    
   


    output.save("Results"+str(filename[:-5])+".xlsx")

